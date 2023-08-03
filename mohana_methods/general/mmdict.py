# -*- coding: utf-8 -*-
"""
Created on Fri Apr  9 16:38:08 2021

@author: mohanam
"""

import openpyxl

def depth(d, init_depth=0):
    """Get the depth of a nested dictionary.
    
    Parameters
    ----------
    d : dict
    init_depth: int
        Default  0
    
    Returns
    -------
    int
        the maximum depth of the dictionary
    """
    if not isinstance(d, dict) or not d:
        return init_depth
    return max(depth(v, init_depth + 1) for k, v in d.items())

def to_excel(d, file_dir, save=True):
    """Save a dictionary to an excel file.
    
    Parameters
    ----------
    d : dict
    file_dir: str
        path to the excel file
    save: bool
        Default True
    
    Returns
    -------
    openpyxl.workbook.workbook.Workbook
    """
    # Create the workbook and sheet for Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Start writting
    _write(d, sheet, 2, 2, depth(d)+2)
    # Save the workbook
    if save:
        workbook.save(filename=file_dir)
    # Return the workbook
    return workbook

def _write(d, sheet, row, col, valcol):
    """Recursive function to write into an 
    openpyxl.worksheet.worksheet.Worksheet a dictionary.
    
    Parameters
    ----------
    d : dict
    sheet : openpyxl.worksheet.worksheet.Worksheet
    row: int
        where to start writting keys
    col: int
        where to start writting keys
    valcol:
        where to start writting values

    Returns
    -------
    int
        where it stopped writting
    """
    # openpyxl does things based on 1 instead of 0
    for key, values in d.items():
        sheet.cell(row=row, column=col, value=key)
        if isinstance(values, dict):
            row = _write(values, sheet, row, col=col+1, valcol=valcol)
        else:
            sheet.cell(row=row, column=valcol, value=str(values))
            row += 1
    return row